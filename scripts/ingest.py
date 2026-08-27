"""Parse the daily inverter generation workbooks into tidy CSVs.

Power Automate drops each morning's report into Generation/MM/DD/'YYYY HH:MM.xlsx'.
That tree is the archive and is never modified here - the workbooks are only read.

The workbook is a formatted report, not a table: values sit in merged cells and
the two sections (PLANT and TOWNSHIP) repeat a Name / CAPACITY / unit / reading
block. Parsing is therefore anchored on those labels rather than on fixed cell
addresses, so a column being inserted or a site being added does not break it.

Some deliveries arrive blank - the plant reading row never populated. Those are
recorded as gaps rather than silently dropped, so a failed report is visible.
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import (  # noqa: E402
    DAILY_CSV, GAPS_CSV, PROCESSED, ROOT, SITES_CSV,
    clean_text, date_from_path, load_config, source_dirs, to_date, to_float,
)

LABEL_COL = 2  # column B carries the row labels and the reading date


def build_anchor_map(ws):
    """Map every merged cell to the top-left cell that actually holds its value."""
    anchors = {}
    for rng in ws.merged_cells.ranges:
        top, left = rng.min_row, rng.min_col
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                anchors[(row, col)] = (top, left)
    return anchors


def cell_value(ws, anchors, row, col):
    row, col = anchors.get((row, col), (row, col))
    return ws.cell(row, col).value


def anchor_columns(ws, anchors, row):
    """Columns in `row` that own a value (merge followers are skipped)."""
    cols = []
    for col in range(LABEL_COL + 1, ws.max_column + 1):
        if anchors.get((row, col), (row, col)) == (row, col):
            cols.append(col)
    return cols


def find_sections(ws, anchors):
    """Locate each section and the Name / CAPACITY / reading rows inside it."""
    sections = []
    for row in range(1, ws.max_row + 1):
        label = clean_text(cell_value(ws, anchors, row, LABEL_COL)).upper()
        if label.startswith("DAILY GENERATION"):
            sections.append({
                "section": "TOWNSHIP" if "TOWNSHIP" in label else "PLANT",
                "header_row": row, "name_row": None, "capacity_row": None,
                "data_row": None, "end_row": ws.max_row,
            })
            if len(sections) > 1:
                sections[-2]["end_row"] = row - 1
            continue
        if not sections:
            continue
        current = sections[-1]
        if label == "NAME":
            current["name_row"] = row
        elif label.startswith("CAPACITY"):
            current["capacity_row"] = row
        elif current["data_row"] is None and to_date(
                cell_value(ws, anchors, row, LABEL_COL)) is not None:
            current["data_row"] = row
    return sections


def locate_readings(ws, anchors, sec):
    """The reading row, by its date; failing that, the first row of numbers.

    A report whose date cell never filled in is still usable if the readings
    landed, so fall back to the shape of the row rather than giving up.
    """
    if sec["data_row"]:
        return sec["data_row"]
    if not sec["capacity_row"]:
        return None
    name_cols = anchor_columns(ws, anchors, sec["name_row"]) if sec["name_row"] else []
    for row in range(sec["capacity_row"] + 1, sec["end_row"] + 1):
        if any(isinstance(cell_value(ws, anchors, row, c), (int, float))
               for c in name_cols):
            return row
    return None


def classify(name: str) -> str:
    upper = name.upper()
    if "TOTAL NET" in upper:
        return "net_total"
    if "TOTAL" in upper:
        return "section_total"
    return "site"


def parse_workbook(path: Path) -> dict:
    """Read one workbook. Never raises for a blank report - reports status instead."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    anchors = build_anchor_map(ws)
    sections = find_sections(ws, anchors)
    rel = path.relative_to(ROOT) if path.is_absolute() else path

    if not sections:
        return {"status": "unreadable", "date": date_from_path(path),
                "reason": "no 'DAILY GENERATION' section headers found",
                "sites": [], "totals": {}, "source": str(rel)}

    sheet_date, rows, totals, missing = None, [], {}, []

    for sec in sections:
        if not (sec["name_row"] and sec["capacity_row"]):
            missing.append(sec["section"])
            continue
        reading_row = locate_readings(ws, anchors, sec)
        if reading_row is None:
            missing.append(sec["section"])
            continue

        found = to_date(cell_value(ws, anchors, reading_row, LABEL_COL))
        if found and sheet_date is None:
            sheet_date = found

        for col in anchor_columns(ws, anchors, sec["name_row"]):
            name = clean_text(cell_value(ws, anchors, sec["name_row"], col))
            if not name:
                continue
            record = {
                "section": sec["section"],
                "column": get_column_letter(col),
                "name": name,
                "capacity_kw": to_float(cell_value(ws, anchors, sec["capacity_row"], col)),
                "energy_kwh": to_float(cell_value(ws, anchors, reading_row, col)),
            }
            kind = classify(name)
            if kind == "site":
                rows.append(record)
            else:
                totals["net" if kind == "net_total" else sec["section"].lower()] = record

    report_date = sheet_date or date_from_path(path)
    plant_rows = [r for r in rows if r["section"] == "PLANT"]
    has_plant = any(r["energy_kwh"] is not None for r in plant_rows)

    if "PLANT" in missing or not has_plant:
        return {"status": "blank", "date": report_date,
                "reason": ("plant reading row never populated"
                           if "PLANT" in missing else "plant readings all empty"),
                "sites": [], "totals": {}, "source": str(rel)}
    if report_date is None:
        return {"status": "unreadable", "date": None,
                "reason": "no date in the sheet and none derivable from the path",
                "sites": [], "totals": {}, "source": str(rel)}

    for r in rows:
        r["date"] = report_date.isoformat()
    return {"status": "ok", "date": report_date, "sites": rows, "totals": totals,
            "dated_in_sheet": sheet_date is not None,
            "partial": [s for s in missing],
            "source": str(rel)}


def summarise(parsed: dict, tolerance: float) -> dict:
    """Reported totals and totals computed from the columns, side by side."""
    sites, totals = parsed["sites"], parsed["totals"]
    out = {"date": parsed["date"].isoformat(), "source_file": parsed["source"]}

    for section in ("PLANT", "TOWNSHIP"):
        members = [s for s in sites if s["section"] == section]
        key = section.lower()
        computed = sum(s["energy_kwh"] or 0.0 for s in members)
        cap_computed = sum(s["capacity_kw"] or 0.0 for s in members)
        reported = totals.get(key, {})
        out[f"{key}_kwh_computed"] = round(computed, 3)
        out[f"{key}_kwh_reported"] = reported.get("energy_kwh")
        out[f"{key}_capacity_kw_computed"] = round(cap_computed, 3)
        out[f"{key}_capacity_kw_reported"] = reported.get("capacity_kw")
        out[f"{key}_site_count"] = len(members)

    net = totals.get("net", {})
    out["net_kwh_reported"] = net.get("energy_kwh")
    out["net_capacity_kw_reported"] = net.get("capacity_kw")
    out["net_kwh_computed"] = round(
        out["plant_kwh_computed"] + out["township_kwh_computed"], 3)
    out["net_capacity_kw_computed"] = round(
        out["plant_capacity_kw_computed"] + out["township_capacity_kw_computed"], 3)

    flags = []
    for key in ("plant", "township", "net"):
        reported, computed = out.get(f"{key}_kwh_reported"), out.get(f"{key}_kwh_computed")
        if reported is not None and computed is not None:
            delta = round(computed - reported, 3)
            out[f"{key}_kwh_delta"] = delta
            if abs(delta) > tolerance:
                flags.append(f"{key}:{delta:+.2f}kWh")
        else:
            out[f"{key}_kwh_delta"] = None
    out["reconciliation_flags"] = ";".join(flags)
    out["dated_in_sheet"] = "yes" if parsed.get("dated_in_sheet") else "no"
    # The column sum is the defensible figure for trending; keep it as the headline.
    out["headline_kwh"] = out["net_kwh_computed"]
    return out


def write_csv(path: Path, rows: list, fields: list):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: ("" if row.get(k) is None else row.get(k))
                             for k in fields})


def discover(config: dict, explicit: list[Path]) -> list[Path]:
    if explicit:
        return explicit
    found = []
    for directory in source_dirs(config):
        if directory.exists():
            found.extend(p for p in directory.rglob("*.xls*")
                         if not p.name.startswith("~$"))
    return sorted(found)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("files", nargs="*", type=Path,
                    help="workbooks to ingest (default: every workbook under the "
                         "configured source directories)")
    ap.add_argument("--quiet", action="store_true", help="only print the summary")
    args = ap.parse_args()

    config = load_config()
    tolerance = config.get("reconciliation_tolerance_kwh", 0.5)
    files = discover(config, args.files)
    if not files:
        print("No workbooks found. Nothing to ingest.")
        return 0

    by_date: dict[str, dict] = {}
    site_rows: dict[str, list] = {}
    gaps: dict[str, dict] = {}

    for path in sorted(files):
        parsed = parse_workbook(path)
        if parsed["status"] != "ok":
            day = parsed["date"].isoformat() if parsed["date"] else "unknown"
            gaps[day] = {"date": day, "status": parsed["status"],
                         "reason": parsed["reason"], "source_file": parsed["source"]}
            if not args.quiet:
                print(f"  GAP  {parsed['source']} -> {day}: {parsed['reason']}")
            continue

        summary = summarise(parsed, tolerance)
        day = summary["date"]
        # A later delivery for the same reading date supersedes an earlier one.
        by_date[day] = summary
        site_rows[day] = parsed["sites"]
        gaps.pop(day, None)
        if not args.quiet:
            note = (f"  [reconcile {summary['reconciliation_flags']}]"
                    if summary["reconciliation_flags"] else "")
            dated = "" if parsed["dated_in_sheet"] else "  [date from path]"
            print(f"  {parsed['source']} -> {day}: "
                  f"{summary['headline_kwh']:>8.2f} kWh{dated}{note}")

    if not by_date:
        print("No workbook parsed successfully.", file=sys.stderr)
        return 1

    PROCESSED.mkdir(parents=True, exist_ok=True)
    flat_sites = [r for day in sorted(site_rows) for r in site_rows[day]]
    write_csv(SITES_CSV, flat_sites,
              ["date", "section", "column", "name", "capacity_kw", "energy_kwh"])

    daily = [by_date[d] for d in sorted(by_date)]
    write_csv(DAILY_CSV, daily, list(daily[0].keys()))
    write_csv(GAPS_CSV, [gaps[d] for d in sorted(gaps)],
              ["date", "status", "reason", "source_file"])

    flagged = sum(1 for r in daily if r["reconciliation_flags"])
    print(f"\n{len(daily)} day(s) ingested, {len(flat_sites)} site-rows, "
          f"{len(gaps)} blank/unreadable report(s), "
          f"{flagged} day(s) with totals that do not reconcile")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
